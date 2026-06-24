"""
Compute deployment metrics and append them to an Excel workbook.

Metric definitions:
  Acc: overall accuracy, trace(cm) / sum(cm)
  F1 : macro one-vs-rest F1 across classes
  FDR: macro detection rate / recall, TP / (TP + FN)
  FPR: macro one-vs-rest false positive rate, FP / (FP + TN)

All metric values are stored as ratios in [0, 1].
"""

import argparse
import json
import os
import re
from collections import defaultdict
from datetime import datetime
from typing import Any, Dict, List, Mapping, Optional, Sequence, Tuple

import numpy as np

from config import Config, SNR_LIST, format_snr


METRIC_NAMES = ('acc', 'f1', 'fdr', 'fpr')
EXCEL_METRIC_COLUMNS = set(METRIC_NAMES)
EXCEL_METRIC_FORMAT = '0.0000'

EXCEL_HEADERS = [
    'dataset',
    'mode',
    'seed',
    'aggregate_kind',
    'domain',
    'n_test',
    'acc',
    'f1',
    'fdr',
    'fpr',
    'timestamp',
    'run_dir',
    'training_mode',
    'single_domain_snr',
    'pu_class_set',
    'progressive_training',
    'epochs',
    'lr',
    'batch_size',
    'target_batch_size',
    'lambda_max',
    'lambda_coral',
    'lambda_sparse',
    'lambda_mmd',
    'patience',
    'ckpt_epoch',
    'ckpt_val_acc',
    'best_epoch',
    'best_val_acc',
]


def _safe_div(num: np.ndarray, den: np.ndarray) -> np.ndarray:
    return np.divide(num, den, out=np.zeros_like(num, dtype=np.float64), where=den != 0)


def _domain_label(domain: Any) -> str:
    return domain if isinstance(domain, str) else format_snr(domain)


def _format_mean_std(mean: float, std: float) -> str:
    return f'{mean:.4f} +/- {std:.4f}'


def _seed_from_run_dir(run_dir: str) -> Optional[int]:
    if not run_dir:
        return None
    base = os.path.basename(os.path.normpath(run_dir))
    m = re.search(r'_seed(-?\d+)$', base)
    if m:
        return int(m.group(1))
    return None


def _metric_ratio(value: Any) -> float:
    return round(float(value), 4)


def _as_ratio(value: Any) -> Any:
    if value is None:
        return None
    try:
        out = float(value)
    except (TypeError, ValueError):
        return value
    if abs(out) > 1.0:
        out /= 100.0
    return round(out, 4)


def metrics_from_confusion_matrix(cm: np.ndarray) -> Dict[str, float]:
    """Return Acc/F1/FDR/FPR from one multiclass confusion matrix."""
    cm = np.asarray(cm, dtype=np.float64)
    total = float(cm.sum())
    tp = np.diag(cm)
    fp = cm.sum(axis=0) - tp
    fn = cm.sum(axis=1) - tp
    tn = total - tp - fp - fn

    precision = _safe_div(tp, tp + fp)
    recall = _safe_div(tp, tp + fn)
    f1 = _safe_div(2.0 * precision * recall, precision + recall)
    fdr = _safe_div(tp, tp + fn)
    fpr = _safe_div(fp, fp + tn)

    return {
        'n_test': int(total),
        'acc': float(_safe_div(np.asarray([tp.sum()]), np.asarray([total]))[0]),
        'f1': float(np.mean(f1)),
        'fdr': float(np.mean(fdr)),
        'fpr': float(np.mean(fpr)),
    }


def compute_metrics_by_domain(cm_by_domain: Mapping[Any, np.ndarray]) -> Dict[Any, Dict[str, float]]:
    return {domain: metrics_from_confusion_matrix(cm) for domain, cm in cm_by_domain.items()}


def average_domain_metrics(metrics_by_domain: Mapping[Any, Mapping[str, float]]) -> Dict[str, float]:
    values = list(metrics_by_domain.values())
    if not values:
        return {'n_test': 0, 'acc': 0.0, 'f1': 0.0, 'fdr': 0.0, 'fpr': 0.0}
    avg = {
        name: float(np.mean([float(m[name]) for m in values]))
        for name in METRIC_NAMES
    }
    avg['n_test'] = int(sum(int(m.get('n_test', 0)) for m in values))
    return avg


def load_confusion_matrices(run_dir: str) -> Dict[Any, np.ndarray]:
    """Load confusion_matrices.npz written by deploy.py."""
    npz_path = os.path.join(run_dir, 'confusion_matrices.npz')
    if not os.path.isfile(npz_path):
        raise FileNotFoundError(f'confusion matrix file not found: {npz_path}')

    npz = np.load(npz_path)
    cm_by_domain: Dict[Any, np.ndarray] = {}
    try:
        for key in npz.files:
            tag = key.replace('domain_', '')
            domain = None if tag == 'clean' else int(tag.replace('dB', ''))
            cm_by_domain[domain] = np.asarray(npz[key], dtype=np.int64)
    finally:
        npz.close()

    ordered = {snr: cm_by_domain[snr] for snr in SNR_LIST if snr in cm_by_domain}
    for domain, cm in cm_by_domain.items():
        if domain not in ordered:
            ordered[domain] = cm
    return ordered


def save_metrics_json(run_dir: str,
                      cfg: Config,
                      metrics_by_domain: Mapping[Any, Mapping[str, float]],
                      extra_meta: Optional[Mapping[str, Any]] = None) -> str:
    avg = average_domain_metrics(metrics_by_domain)
    path = os.path.join(run_dir, 'deploy_metrics.json')
    payload = {
        'mode': cfg.mode,
        'dataset': cfg.dataset,
        'seed': cfg.seed,
        'metric_unit': 'ratio',
        'metric_definition': {
            'acc': 'overall accuracy',
            'f1': 'macro one-vs-rest F1',
            'fdr': 'macro detection rate / recall',
            'fpr': 'macro one-vs-rest false positive rate',
        },
        'per_domain': {
            _domain_label(domain): dict(metrics)
            for domain, metrics in metrics_by_domain.items()
        },
        'average': avg,
    }
    if extra_meta:
        payload['extra'] = dict(extra_meta)
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    return path


def _read_json(path: str) -> Dict[str, Any]:
    if not os.path.isfile(path):
        return {}
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


def _read_saved_deploy_meta(run_dir: str) -> Dict[str, Any]:
    deploy_json = _read_json(os.path.join(run_dir, 'deploy_results.json'))
    return {
        'ckpt_epoch': deploy_json.get('ckpt_epoch'),
        'ckpt_val_acc': deploy_json.get('ckpt_val_acc'),
    }


def _run_metadata(cfg: Config,
                  run_dir: str,
                  extra_meta: Optional[Mapping[str, Any]] = None) -> Dict[str, Any]:
    cfg_json = _read_json(os.path.join(run_dir, 'config.json'))
    history_json = _read_json(os.path.join(run_dir, 'train_history.json'))
    deploy_meta = _read_saved_deploy_meta(run_dir)

    def pick(name: str, default: Any = None) -> Any:
        return cfg_json.get(name, getattr(cfg, name, default))

    meta = {
        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'run_dir': run_dir,
        'dataset': pick('dataset', cfg.dataset),
        'mode': pick('mode', cfg.mode),
        'seed': pick('seed', cfg.seed),
        'training_mode': pick('training_mode'),
        'single_domain_snr': pick('single_domain_snr'),
        'pu_class_set': pick('pu_class_set'),
        'progressive_training': pick('progressive_training'),
        'epochs': pick('epochs'),
        'lr': pick('lr'),
        'batch_size': pick('batch_size'),
        'target_batch_size': pick('target_batch_size'),
        'lambda_max': pick('lambda_max'),
        'lambda_coral': pick('lambda_coral'),
        'lambda_sparse': pick('lambda_sparse'),
        'lambda_mmd': pick('lambda_mmd'),
        'patience': pick('patience'),
        'ckpt_epoch': deploy_meta.get('ckpt_epoch'),
        'ckpt_val_acc': _as_ratio(deploy_meta.get('ckpt_val_acc')),
        'best_epoch': history_json.get('best_epoch'),
        'best_val_acc': _as_ratio(history_json.get('best_val_acc')),
        'aggregate_kind': 'seed',
    }
    if extra_meta:
        meta.update(extra_meta)
    return meta


def build_excel_rows(cfg: Config,
                     metrics_by_domain: Mapping[Any, Mapping[str, float]],
                     run_dir: Optional[str] = None,
                     extra_meta: Optional[Mapping[str, Any]] = None) -> List[Dict[str, Any]]:
    run_dir = run_dir or cfg.run_dir()
    meta = _run_metadata(cfg, run_dir, extra_meta=extra_meta)
    rows: List[Dict[str, Any]] = []

    for domain, metrics in metrics_by_domain.items():
        row = dict(meta)
        row.update({
            'domain': _domain_label(domain),
            'n_test': int(metrics.get('n_test', 0)),
            'acc': _metric_ratio(metrics.get('acc', 0.0)),
            'f1': _metric_ratio(metrics.get('f1', 0.0)),
            'fdr': _metric_ratio(metrics.get('fdr', 0.0)),
            'fpr': _metric_ratio(metrics.get('fpr', 0.0)),
        })
        rows.append(row)

    avg = average_domain_metrics(metrics_by_domain)
    row = dict(meta)
    row.update({
        'domain': 'AVG',
        'n_test': int(avg.get('n_test', 0)),
        'acc': _metric_ratio(avg.get('acc', 0.0)),
        'f1': _metric_ratio(avg.get('f1', 0.0)),
        'fdr': _metric_ratio(avg.get('fdr', 0.0)),
        'fpr': _metric_ratio(avg.get('fpr', 0.0)),
    })
    rows.append(row)
    return rows


def build_seed_aggregate_rows(rows_by_seed: Sequence[Sequence[Mapping[str, Any]]]) -> List[Dict[str, Any]]:
    flat: List[Dict[str, Any]] = []
    for rows in rows_by_seed:
        flat.extend(dict(r) for r in rows)

    if not flat:
        return []

    grouped: Dict[Tuple[Any, Any, Any], List[Dict[str, Any]]] = defaultdict(list)
    for row in flat:
        if str(row.get('aggregate_kind', 'seed')) != 'seed':
            continue
        key = (row.get('dataset'), row.get('mode'), row.get('domain'))
        grouped[key].append(row)

    summary_rows: List[Dict[str, Any]] = []
    for (dataset, mode, domain), group in grouped.items():
        base = dict(group[0])
        for metric in METRIC_NAMES:
            values = [float(r.get(metric, 0.0) or 0.0) for r in group]
            mean = float(np.mean(values))
            std = float(np.std(values, ddof=0))
            base[metric] = _format_mean_std(mean, std)
        base['aggregate_kind'] = 'MEAN+/-STD'
        base['seed'] = 'MEAN+/-STD'
        base['run_dir'] = ''
        base['timestamp'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        base['n_test'] = int(np.mean([int(r.get('n_test', 0) or 0) for r in group]))
        summary_rows.append(base)

    return summary_rows


def append_metrics_to_excel(cfg: Config,
                            metrics_by_domain: Mapping[Any, Mapping[str, float]],
                            excel_path: str,
                            sheet_name: str = 'deploy_metrics',
                            run_dir: Optional[str] = None,
                            extra_meta: Optional[Mapping[str, Any]] = None) -> str:
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError(
            'openpyxl is required to write Excel files. Install it with: pip install openpyxl'
        ) from exc

    rows = build_excel_rows(cfg, metrics_by_domain, run_dir=run_dir, extra_meta=extra_meta)
    parent = os.path.dirname(os.path.abspath(excel_path))
    if parent:
        os.makedirs(parent, exist_ok=True)

    if os.path.isfile(excel_path):
        wb = load_workbook(excel_path)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

    # A fresh openpyxl workbook starts with one empty row. Older interrupted runs may
    # also have written headers on row 2 after that blank row; normalize both cases.
    if ws.max_row >= 2 and all(cell.value is None for cell in ws[1]):
        second_row = [cell.value for cell in ws[2] if cell.value]
        if second_row:
            ws.delete_rows(1)

    raw_existing_headers = []
    if ws.max_row >= 1:
        raw_existing_headers = [cell.value for cell in ws[1]]
    existing_headers = [header for header in raw_existing_headers if header]

    if not existing_headers:
        headers = list(EXCEL_HEADERS)
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
    else:
        headers = list(EXCEL_HEADERS)
        for header in existing_headers:
            if header not in headers:
                headers.append(header)

        if existing_headers != headers:
            old_rows = []
            for values in ws.iter_rows(min_row=2, max_col=len(raw_existing_headers), values_only=True):
                if all(value is None for value in values):
                    continue
                old_row = {}
                for idx, header in enumerate(raw_existing_headers):
                    if header and header not in old_row:
                        value = values[idx] if idx < len(values) else None
                        old_row[header] = value
                old_rows.append(old_row)

            ws.delete_rows(1, ws.max_row)
            for col, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col, value=header)
            for old_row in old_rows:
                ws.append([old_row.get(header) for header in headers])

    for row in rows:
        ws.append([row.get(header) for header in headers])

    for col, header in enumerate(headers, start=1):
        if header in EXCEL_METRIC_COLUMNS:
            col_letter = get_column_letter(col)
            for cell in ws[col_letter][1:]:
                cell.number_format = EXCEL_METRIC_FORMAT

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    wb.save(excel_path)
    return excel_path


def append_seed_aggregate_to_excel(cfg: Config,
                                   rows_by_seed: Sequence[Sequence[Mapping[str, Any]]],
                                   excel_path: str,
                                   sheet_name: str = 'deploy_metrics',
                                   include_flat_rows: bool = True) -> str:
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError(
            'openpyxl is required to write Excel files. Install it with: pip install openpyxl'
        ) from exc

    flat_rows = [dict(row) for rows in rows_by_seed for row in rows]
    summary_rows = build_seed_aggregate_rows(rows_by_seed)
    parent = os.path.dirname(os.path.abspath(excel_path))
    if parent:
        os.makedirs(parent, exist_ok=True)

    if os.path.isfile(excel_path):
        wb = load_workbook(excel_path)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

    if ws.max_row >= 2 and all(cell.value is None for cell in ws[1]):
        second_row = [cell.value for cell in ws[2] if cell.value]
        if second_row:
            ws.delete_rows(1)

    headers = list(EXCEL_HEADERS)
    for row in flat_rows + summary_rows:
        for key in row.keys():
            if key not in headers:
                headers.append(key)

    existing_headers = [cell.value for cell in ws[1] if cell.value] if ws.max_row >= 1 else []
    if existing_headers != headers:
        # Rebuild sheet to match stable header order, preserving existing rows.
        old_rows = []
        if existing_headers:
            for values in ws.iter_rows(min_row=2, max_col=len(existing_headers), values_only=True):
                if all(value is None for value in values):
                    continue
                old_rows.append({existing_headers[i]: values[i] for i in range(len(existing_headers))})
        ws.delete_rows(1, ws.max_row)
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
        for old_row in old_rows:
            ws.append([old_row.get(header) for header in headers])
    elif not existing_headers:
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)

    if include_flat_rows:
        for row in flat_rows:
            ws.append([row.get(header) for header in headers])
    for row in summary_rows:
        ws.append([row.get(header) for header in headers])

    for col, header in enumerate(headers, start=1):
        if header in EXCEL_METRIC_COLUMNS:
            col_letter = get_column_letter(col)
            for cell in ws[col_letter][1:]:
                cell.number_format = EXCEL_METRIC_FORMAT

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    wb.save(excel_path)
    return excel_path


def collect_rows_from_mode(results_root: str,
                           dataset: str,
                           mode: str) -> List[List[Dict[str, Any]]]:
    rows_by_seed: List[List[Dict[str, Any]]] = []
    for run_dir, metrics_by_domain in collect_metrics_from_mode(results_root, dataset, mode):
        cfg_data = _read_json(os.path.join(run_dir, 'config.json'))
        fields = Config.__dataclass_fields__
        if cfg_data:
            cfg = Config(**{k: v for k, v in cfg_data.items() if k in fields})
        else:
            cfg = Config(mode=mode, dataset=dataset)
        rows_by_seed.append(build_excel_rows(cfg, metrics_by_domain, run_dir=run_dir))
    return rows_by_seed


def append_mode_aggregate_to_excel(results_root: str,
                                   dataset: str,
                                   mode: str,
                                   excel_path: str,
                                   sheet_name: str = 'deploy_metrics',
                                   include_flat_rows: bool = True) -> str:
    cfg = Config(mode=mode, dataset=dataset)
    rows_by_seed = collect_rows_from_mode(results_root, dataset, mode)
    if not rows_by_seed:
        return excel_path
    return append_seed_aggregate_to_excel(
        cfg,
        rows_by_seed,
        excel_path,
        sheet_name=sheet_name,
        include_flat_rows=include_flat_rows,
    )


def metrics_from_existing_run(cfg: Config,
                              excel_path: Optional[str] = None,
                              sheet_name: str = 'deploy_metrics',
                              run_dir: Optional[str] = None) -> Dict[Any, Dict[str, float]]:
    run_dir = run_dir or cfg.run_dir()
    cm_by_domain = load_confusion_matrices(run_dir)
    metrics_by_domain = compute_metrics_by_domain(cm_by_domain)
    save_metrics_json(run_dir, cfg, metrics_by_domain)
    if excel_path:
        append_metrics_to_excel(cfg, metrics_by_domain, excel_path, sheet_name, run_dir=run_dir)
    return metrics_by_domain


def _iter_run_dirs_for_mode(results_root: str, dataset: str, mode: str) -> List[str]:
    if not results_root or not os.path.isdir(results_root):
        return []
    out = []
    pattern = f'_{dataset}_{mode}_seed'
    for name in os.listdir(results_root):
        path = os.path.join(results_root, name)
        if not os.path.isdir(path):
            continue
        if pattern in name and os.path.isfile(os.path.join(path, 'config.json')):
            out.append(path)
    def _sort_key(path: str):
        seed = _seed_from_run_dir(path)
        return (seed is None, seed if seed is not None else 10**9, path)
    return sorted(out, key=_sort_key)


def collect_metrics_from_mode(results_root: str,
                              dataset: str,
                              mode: str) -> List[Tuple[str, Dict[Any, Dict[str, float]]]]:
    collected: List[Tuple[str, Dict[Any, Dict[str, float]]]] = []
    for run_dir in _iter_run_dirs_for_mode(results_root, dataset, mode):
        npz_path = os.path.join(run_dir, 'confusion_matrices.npz')
        if not os.path.isfile(npz_path):
            print(f'[WARN] skip {run_dir}: confusion_matrices.npz not found, run deploy first')
            continue
        cfg_path = os.path.join(run_dir, 'config.json')
        cfg_data = _read_json(cfg_path)
        fields = Config.__dataclass_fields__
        cfg = Config(**{k: v for k, v in cfg_data.items() if k in fields}) if cfg_data else Config(mode=mode, dataset=dataset)
        cm_by_domain = load_confusion_matrices(run_dir)
        metrics_by_domain = compute_metrics_by_domain(cm_by_domain)
        save_metrics_json(run_dir, cfg, metrics_by_domain)
        collected.append((run_dir, metrics_by_domain))
    return collected


def _print_metrics(metrics_by_domain: Mapping[Any, Mapping[str, float]]) -> None:
    print('domain    n_test        acc         f1        fdr        fpr')
    print('-' * 65)
    for domain, metrics in metrics_by_domain.items():
        print(f"{_domain_label(domain):>6s}  {int(metrics['n_test']):7d}  "
              f"{metrics['acc']:9.4f}  {metrics['f1']:9.4f}  "
              f"{metrics['fdr']:9.4f}  {metrics['fpr']:9.4f}")
    avg = average_domain_metrics(metrics_by_domain)
    print('-' * 65)
    print(f"{'AVG':>6s}  {int(avg['n_test']):7d}  "
          f"{avg['acc']:9.4f}  {avg['f1']:9.4f}  "
          f"{avg['fdr']:9.4f}  {avg['fpr']:9.4f}")


def parse_args():
    p = argparse.ArgumentParser('Deploy a trained model and export Acc/F1/FDR/FPR metrics')
    p.add_argument('--mode', type=str, default='E1_QMDCS')
    p.add_argument('--dataset', type=str, default='PU', choices=['PU', 'CWRU'])
    p.add_argument('--seed', type=int, default=42)
    p.add_argument('--results_root', type=str, default=None)
    p.add_argument('--no_cuda', action='store_true')
    p.add_argument('--pu_class_set', type=str, default=None, choices=['4class', '10class'])
    p.add_argument('--excel', type=str, default=None,
                   help='Excel workbook to append metrics into, for example results/deploy_metrics.xlsx')
    p.add_argument('--excel_sheet', type=str, default='deploy_metrics')
    p.add_argument('--from_npz', action='store_true',
                   help='Only compute metrics from an existing confusion_matrices.npz; do not run deploy.')
    p.add_argument('--aggregate_existing', action='store_true',
                   help='Scan existing run dirs under results_root for the given dataset/mode and append seed rows plus mean/std rows to Excel only.')
    p.add_argument('--aggregate_summary_only', action='store_true',
                   help='With --aggregate_existing, append only mean/std rows; use after seed rows were already written.')
    p.add_argument('--run_dir', type=str, default=None,
                   help='Existing run directory. With --from_npz, read confusion_matrices.npz; otherwise deploy checkpoint.pth here.')
    p.add_argument('--ckpt', type=str, default=None,
                   help='Trained checkpoint path to deploy. Defaults to <run_dir>/checkpoint.pth or cfg.ckpt_path().')
    return p.parse_args()


def cfg_from_args(args) -> Config:
    cfg_data = None
    if args.run_dir:
        config_path = os.path.join(args.run_dir, 'config.json')
        if os.path.isfile(config_path):
            cfg_data = _read_json(config_path)

    if cfg_data:
        fields = Config.__dataclass_fields__
        cfg = Config(**{k: v for k, v in cfg_data.items() if k in fields})
    else:
        cfg = Config(mode=args.mode, dataset=args.dataset, seed=args.seed)

    if args.results_root is not None:
        cfg.results_root = args.results_root
    cfg.no_cuda = args.no_cuda
    if args.pu_class_set is not None:
        cfg.pu_class_set = args.pu_class_set
    return cfg


def main():
    args = parse_args()
    cfg = cfg_from_args(args)

    if args.aggregate_existing:
        if not args.excel:
            raise ValueError('--aggregate_existing requires --excel')
        results_root = args.results_root or cfg.results_root
        out = append_mode_aggregate_to_excel(
            results_root=results_root,
            dataset=cfg.dataset,
            mode=cfg.mode,
            excel_path=args.excel,
            sheet_name=args.excel_sheet,
            include_flat_rows=not args.aggregate_summary_only,
        )
        print(f'[EXCEL] appended mode aggregate rows to {out} (dataset={cfg.dataset}, mode={cfg.mode})')
        return

    if args.from_npz:
        metrics = metrics_from_existing_run(
            cfg,
            excel_path=args.excel,
            sheet_name=args.excel_sheet,
            run_dir=args.run_dir,
        )
        _print_metrics(metrics)
        if args.excel:
            print(f'[EXCEL] appended metrics to {args.excel}')
        return

    from deploy import run_deploy

    run_deploy(cfg, excel_path=args.excel, excel_sheet=args.excel_sheet,
               run_dir=args.run_dir, ckpt_path=args.ckpt)


if __name__ == '__main__':
    main()
